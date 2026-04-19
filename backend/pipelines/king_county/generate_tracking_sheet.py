"""
SellerSignal v3 — King County generate_tracking_sheet orchestrator (v2 reference port).

Produces a companion XLSX workbook for agent outcome tracking. Two sheets:
  Sheet 1 'This Week's Plays' — 10 picks pre-filled with status dropdown
  Sheet 2 'How to Use' — instructions

Agents update Status, Contact Date, Response, Next Step, Notes over the
week, then return the sheet. Future: ingest_outcomes.py merges back.

STATUS: NOT RUNNABLE IN V3 AS-IS.
__main__ block hardcodes v2 sandbox paths. The build_workbook()
function itself is parameterized and reusable.

DEPENDENCIES:
  - openpyxl (external Python package)

CONTAINS:
  - KC-specific: ZIP_TO_CITY dict (6 Eastside ZIPs), __main__ block
    with v2 sandbox paths.
  - Parameterized: build_workbook(picks_path, out_path) — callable
    from any orchestrator.

PURPOSE:
  1. Specification — documents the agent tracking XLSX format.
  2. Migration target — when v3's KC orchestrator wires this into the
     Supabase-backed flow, this file is the diff target.

Original v2 docstring:

  generate_tracking_sheet.py — Companion XLSX for agent outcome tracking.

  Reads this-weeks-picks.json, produces an Excel workbook with:
    Sheet 1 "This Week's Plays" — pre-filled with the 10 picks + status dropdown
    Sheet 2 "How to Use" — brief instructions

  Agent fills in Status, Contact Date, Response, Next Step, Notes over the week,
  then hands back. (Future: ingest_outcomes.py merges back into outcomes.json.)
"""
import json, os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


STATUS_OPTIONS = ['NEW', 'CONTACTED', 'RESPONDED', 'MEETING', 'LISTING', 'LOST', 'DEAD']

# Palette tuned for Excel — the Estate aesthetic
IVORY_HEX = 'F7F1E8'
IVORY_DEEP = 'EDE4D3'
INK_HEX = '1A1A1A'
INK_SOFT = '3C3C3A'
GOLD_HEX = 'A88D4A'
GOLD_LIGHT = 'D4BC82'
RED_HEX = '8C3B36'

ZIP_TO_CITY = {'98004': 'Bellevue', '98039': 'Medina', '98040': 'Mercer Island',
               '98033': 'Kirkland', '98006': 'Newport', '98005': 'Bridle Trails'}


def build_workbook(picks_path, out_path):
    picks = json.load(open(picks_path))
    week_of_dt = datetime.strptime(picks['week_of'], '%Y-%m-%d')
    week_of_str = week_of_dt.strftime('%B %-d, %Y')

    wb = Workbook()
    ws = wb.active
    ws.title = "This Week's Plays"

    # Global default font
    default_font = Font(name='Arial', size=10, color=INK_HEX)

    # ─── TITLE ROWS ──────────────────────────────────────────────────
    ws['A1'] = 'SELLERSIGNAL  ·  OPERATOR TRACKING'
    ws['A1'].font = Font(name='Arial', size=9, color=GOLD_HEX, bold=True)
    ws['A2'] = f"Week of {week_of_str}"
    ws['A2'].font = Font(name='Arial', size=16, color=INK_HEX, bold=True)
    ws['A3'] = 'Ten moves. Five to call, three to cultivate, two long-setup plays.'
    ws['A3'].font = Font(name='Arial', size=10, color=INK_SOFT, italic=True)

    # ─── HEADER ROW ──────────────────────────────────────────────────
    headers = ['Section', 'Address', 'City', 'Owner', 'Value', 'Signal',
               'Status', 'Contact Date', 'Response', 'Next Step', 'Notes']
    header_row = 5
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = Font(name='Arial', size=9, color='FFFFFF', bold=True)
        c.fill = PatternFill('solid', start_color=GOLD_HEX)
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border = Border(bottom=Side(border_style='medium', color=INK_HEX))

    # ─── DATA ROWS ───────────────────────────────────────────────────
    section_colors = {
        'CALL NOW':        RED_HEX,
        'BUILD NOW':       GOLD_HEX,
        'STRATEGIC HOLDS': INK_SOFT,
    }
    row = header_row + 1
    first_status_row = row

    all_moves = []
    for p in picks['call_now']:        all_moves.append(('CALL NOW', p))
    for p in picks['build_now']:       all_moves.append(('BUILD NOW', p))
    for p in picks['strategic_holds']: all_moves.append(('STRATEGIC HOLDS', p))

    for section, p in all_moves:
        city = ZIP_TO_CITY.get(p.get('zip'), '')
        sig = (p.get('signal_family') or '').replace('_', ' ').title()
        # Section tag
        c = ws.cell(row=row, column=1, value=section)
        c.font = Font(name='Arial', size=9, color='FFFFFF', bold=True)
        c.fill = PatternFill('solid', start_color=section_colors.get(section, GOLD_HEX))
        c.alignment = Alignment(horizontal='left', vertical='center')
        # Body cells
        ws.cell(row=row, column=2, value=p.get('address', '—')).font = Font(name='Arial', size=10, color=INK_HEX, bold=True)
        ws.cell(row=row, column=3, value=city).font = default_font
        ws.cell(row=row, column=4, value=(p.get('owner') or '—')[:60]).font = default_font
        val = p.get('value', 0) or 0
        v_cell = ws.cell(row=row, column=5, value=val)
        v_cell.number_format = '"$"#,##0'
        v_cell.font = default_font
        v_cell.alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=6, value=sig).font = default_font
        # Status — blank by default, dropdown validation added below
        status_cell = ws.cell(row=row, column=7, value='NEW')
        status_cell.font = Font(name='Arial', size=10, color=INK_HEX, bold=True)
        status_cell.fill = PatternFill('solid', start_color=IVORY_DEEP)
        status_cell.alignment = Alignment(horizontal='center', vertical='center')
        # Editable columns - faint ivory fill to indicate agent input
        for col in range(8, 12):
            cell = ws.cell(row=row, column=col, value='')
            cell.fill = PatternFill('solid', start_color=IVORY_HEX)
            cell.font = default_font
        # Row shading for section readability
        if section == 'CALL NOW':
            for col in range(2, 12):
                ws.cell(row=row, column=col).border = Border(
                    bottom=Side(border_style='thin', color=IVORY_DEEP))
        row += 1

    last_status_row = row - 1

    # ─── DATA VALIDATION for Status column ───────────────────────────
    dv = DataValidation(type='list',
                        formula1=f'"{",".join(STATUS_OPTIONS)}"',
                        allow_blank=False)
    dv.add(f'G{first_status_row}:G{last_status_row}')
    dv.error = 'Status must be one of: ' + ', '.join(STATUS_OPTIONS)
    dv.errorTitle = 'Invalid status'
    dv.prompt = 'NEW · CONTACTED · RESPONDED · MEETING · LISTING · LOST · DEAD'
    dv.promptTitle = 'Status'
    ws.add_data_validation(dv)

    # Date validation on Contact Date
    date_dv = DataValidation(type='date', allow_blank=True)
    date_dv.add(f'H{first_status_row}:H{last_status_row}')
    date_dv.prompt = 'YYYY-MM-DD'
    ws.add_data_validation(date_dv)

    # ─── COLUMN WIDTHS ───────────────────────────────────────────────
    widths = {'A': 16, 'B': 28, 'C': 14, 'D': 28, 'E': 12, 'F': 18,
              'G': 13, 'H': 14, 'I': 28, 'J': 28, 'K': 32}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    # Row heights
    ws.row_dimensions[header_row].height = 22
    for r in range(first_status_row, last_status_row + 1):
        ws.row_dimensions[r].height = 20

    # Freeze header
    ws.freeze_panes = f'A{first_status_row}'

    # ─── HOW-TO SHEET ────────────────────────────────────────────────
    ws2 = wb.create_sheet('How to Use')
    ws2['A1'] = 'How to Use This Sheet'
    ws2['A1'].font = Font(name='Arial', size=16, color=INK_HEX, bold=True)

    rows_text = [
        ('',         ''),
        ('Status',   'Update weekly. NEW → CONTACTED (letter/call sent) → RESPONDED (they replied) → MEETING (scheduled) → LISTING (signed) or LOST (pursued and declined) or DEAD (no interest, do not resurface).'),
        ('',         ''),
        ('DEAD',     'Marks the lead as permanently excluded. Use when the owner has explicitly said no, or clearly will not transact.'),
        ('LOST',     'Temporarily excluded for 90 days. Use when the pursuit failed but the property might re-enter the pipeline later.'),
        ('LISTING',  'Marks as converted. These come out of the weekly rotation.'),
        ('',         ''),
        ('Contact Date', 'Date of first outreach. Leave blank until you contact.'),
        ('Response',     'One-line summary of what happened. Used for pattern analysis.'),
        ('Next Step',    'Your planned follow-up. Helps with continuity week-to-week.'),
        ('Notes',        'Free-form. Anything relevant — referral source, neighbor context, local info.'),
        ('',             ''),
        ('Return',   'Save the workbook and send back by end of week. System ingests outcomes to inform next week\'s selection.'),
        ('',         ''),
        ('Cadence',  'A new sheet arrives every Monday. Same format, different picks. Leads marked DEAD never resurface. Leads marked NEW/CONTACTED/RESPONDED can resurface after a 4-week cooldown if still in inventory.'),
    ]
    for i, (label, text) in enumerate(rows_text, 3):
        c1 = ws2.cell(row=i, column=1, value=label)
        c2 = ws2.cell(row=i, column=2, value=text)
        c1.font = Font(name='Arial', size=10, color=GOLD_HEX, bold=True)
        c2.font = Font(name='Arial', size=10, color=INK_HEX)
        c2.alignment = Alignment(wrap_text=True, vertical='top')
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 90

    wb.save(out_path)
    return out_path, os.path.getsize(out_path)


if __name__ == "__main__":
    path, size = build_workbook(
        '/home/claude/sellersignal_v2/out/this-weeks-picks.json',
        '/home/claude/sellersignal_v2/out/this-weeks-tracking.xlsx',
    )
    print(f"✓ {path}  ({size:,} bytes)")
